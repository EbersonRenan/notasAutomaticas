from playwright.sync_api import Playwright, sync_playwright, expect
import time
import datetime
from openpyxl import load_workbook , workbook
import openpyxl
import os

def run(playwright: Playwright) -> None:
    
    browser = playwright.chromium.launch(channel="chrome", headless=False)
    context = browser.new_context()
    page = context.new_page()
    page.goto("https://nfse-doisirmaos.atende.net/autoatendimento/servicos/nfse?redirected=1")
    page.get_by_placeholder("CPF ou CNPJ").click()
    page.get_by_placeholder("CPF ou CNPJ").fill("02.276.965/0001-37")
    page.get_by_placeholder("Senha").click()
    page.get_by_placeholder("Senha").fill("Mariaclara1582*")
    page.get_by_role("button", name="Entrar", exact=True).click()
    page.get_by_role("link", name="Acessar").click()
    time.sleep(30)
    #page.frame_locator("iframe[name=\"a-b0lkgxhwcgia\"]").get_by_label("Não sou um robô").click()
    
    with page.expect_popup() as page1_info:

        #page.frame_locator("iframe[name=\"c-b0lkgxhwcgia\"]").get_by_role("button", name="Verificar").click()
        page1 = page1_info.value  
        #time.sleep(60)

        def verificar_seletor(page1):
            try:
                page1.wait_for_selector("div.estrutura_conjunto_container_icone")
                return True
            except:
                return False 
        while verificar_seletor(page1)  == False :
            print('aguardando...')
            if verificar_seletor(page1):
                print("esta caregado")
                time.sleep(10)
            else:
                print("nao esta carregado")
                time.sleep(60)

        
        script_folder = os.path.abspath(os.path.dirname(__file__))
        
        print("Script folder:", script_folder)        
        
        
        # conecta na planilha
        workbook = openpyxl.load_workbook('C:\\python\\ÁreadeTrabalho\\nativa\\Pasta1.xlsx')  #Tem que estar na mesma pasta
        
        planilha = workbook['Planilha1']
        #planilha = workbook.active
        
        # Abre a aba.
        sheet = workbook.get_sheet_by_name('Planilha1')

        planilha = sheet
        # Obtenha os valores da coluna A.
        for row in range(2, planilha.max_row + 1):
    
            #Coluna Matr. - Nome do Aluno
            celula = planilha.cell(row=row, column=1)
            MatrNomedoAluno1 = celula.value
            
            #COLUNA CNPJ/CPF
            celula = planilha.cell(row=row, column=2)
            CNPJ_CPF = celula.value
            #tratamwnro de dados
            CNPJ_CPF = CNPJ_CPF.replace(".","")
            CNPJ_CPF = CNPJ_CPF.replace("-","")


            #Coluna Mes
            celula = planilha.cell(row=row, column=3)
            ValorMes = celula.value
            #tratamwnro de dados
            ValorMes = str(ValorMes)
            ValorMes = ValorMes.replace(".",",")
            

            #coluna Status
            celula = planilha.cell(row=row, column=4)
            status = celula.value
               
            #Variaveis:
            cpf_cliente = list(CNPJ_CPF)
            dataAtual= datetime.datetime.now()

            #Executa a emição da nota se o status for diferente de "ok"
            if status != "ok":
                
                try:                
                    page1.get_by_label("Conjuntos").locator("div").filter(has_text="Visão Gerencial").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Visões Gerenciais Disponíveis").locator("div").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_text("Emitir Nota Fiscal").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_role("button", name="Próximo").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_placeholder("Pesquisar por CPF").click()
                    page1.wait_for_timeout(1000) 

                    #Digita o CNPJ
                    page1.get_by_placeholder("Pesquisar por CPF").click()
                    for char in cpf_cliente:
                        page1.keyboard.type(char)
                        page1.wait_for_timeout(1000)

                    page1.wait_for_timeout(3000)
                    page1.get_by_placeholder("Pesquisar por CPF").press("Enter")
                    page1.wait_for_timeout(1000)
                    page1.get_by_role("button", name="Próximo").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Local da Prestação").fill("8625")
                    page1.wait_for_timeout(1000)
                    page1.get_by_role("textbox", name="Digite aqui para consultar").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Valor do Serviço").click() 
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Valor do Serviço").fill(ValorMes)
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Descrição").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Descrição").press("CapsLock")
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Descrição").fill("ACADEMIA")
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Lista de Serviço").select_option("604")
                    page1.wait_for_timeout(1000)
                    page1.get_by_role("button", name="Próximo").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_text("Inf. Complementares").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Imprimir após confirmação?").uncheck()
                    page1.wait_for_timeout(1000)
                    page1.get_by_role("button", name="Emitir").click()
                    page1.wait_for_timeout(2000)                

                    #Da um "ok"
                    celula = planilha.cell(row=row, column=4)
                    celula.value = 'ok'
                    
                    workbook.save("C:\\python\\ÁreadeTrabalho\\nativa\\Pasta1.xlsx")


                except:
                   #Indica erro na planilha
                    celula = planilha.cell(row=row, column=4)
                    celula.value = 'Erro! Por favor verifique o cadastro'
                    
                    workbook.save("C:\\python\\ÁreadeTrabalho\\nativa\\Pasta1.xlsx")
     

    # ---------------------
    context.close()
    browser.close()

with sync_playwright() as playwright:
    run(playwright)